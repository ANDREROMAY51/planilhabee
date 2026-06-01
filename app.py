"""Aplicativo Streamlit para unir planilhas extraídas de um arquivo ZIP."""

import os
import time
import streamlit as st
from xls2xlsx import XLS2XLSX
from zipfile import ZipFile
import shutil
import re
from openpyxl import load_workbook
import copy

pasta_atual = os.getcwd()
pasta_descompactados = f'{pasta_atual}\\Arquivos Descompactados\\'
planilha_principal = 'PLANILHAFOLHA.xlsx'
nome_plan_original = 'PLANILHAFOLHA'
path_planilha_principal = f'{pasta_descompactados}{planilha_principal}'

st.set_page_config(page_title = "Planilhas Bee", layout = "wide")
st.title("App para juntar as planilhas") 
tempo_alerta = st.empty() # para ficar 3 segundos de alerta e apagar  

# UNZIP FILES-----------------------------------------    
uploaded_file = st.file_uploader("Por favor, escolha a pasta para descompactar", type="zip")
if uploaded_file is not None:
    with ZipFile(uploaded_file, 'r') as z:
        # List files in the ZIP
        filenames = z.namelist()       
        z.extractall(pasta_descompactados)                                  
    
    with st.status("Iniciando processo...", expanded=True) as status:
        time.sleep(3)
        st.write("Descompactando arquivos...")
        time.sleep(3)
        st.write("Processando arquivos...")
        time.sleep(5)
        status.update(label="Processo concluído, aguardando liberação da planilha...", state="complete")     
                
    if pasta_descompactados:       
        for plan in os.listdir(pasta_descompactados):
            planilha = os.path.join(pasta_descompactados, plan)
            #EEXCLUI OS ARQUIVOS XLSX PARA GERAR NOVAMENTE PARA NÃO DUPLICAR
            if planilha.endswith(".xlsx"):
                os.remove(planilha)
                
            planilha = planilha.replace("/", "\\")
            #planilhas_para_importar.append(planilha)        
                                
            # Lê o arquivo XLS para transformar em XLSX                 
            try:
                x2x = XLS2XLSX(planilha)
            except:
                continue
            # Save as .xlsx
            planilha_xls = planilha.replace("Planilha_Folha_","")
            x2x.to_xlsx(f'{planilha_xls}.xlsx')
            
            #FIANLIZA REMOVENDO OS ARQUIVOS XLS
            if planilha.endswith(".xls"):
                os.remove(planilha)
                
            for arq in os.listdir(pasta_descompactados):
                if "412" in arq:
                    arquivo = os.path.join(pasta_descompactados, arq) 
                    principal = os.path.join(pasta_descompactados, planilha_principal)                
                    #renome o arquivo
                    shutil.copy(arquivo, principal)  
                    
        #CRIA A LISTA DOS NOMES DAS PLANILHAS---------------------------------------
        lista_planilhas = []
        for arq in os.listdir(pasta_descompactados):
            if "412" in arq: 
                #remove o arquivo 412 xlsx
                arq_412 = os.path.join(pasta_descompactados, arq)
                os.remove(arq_412)
                continue
            else:
                #ajustes no nome do arquivo
                plan = re.sub(r'\d+', '', arq).replace(".xlsx", "").replace(".xls", "")
                plan = re.sub(r"VOLARE", '', plan)
                plan = re.sub(r"SALUTE", '', plan)
                plan = re.sub(r"\(|\)", '', plan) #elimina parenteses                
                plan = re.sub(r"_", '', plan)
                plan = re.sub(r"-", '', plan)
                plan = plan.replace(" ", "") # elimina os espaços em branco
                #Limita o tamanho a aba até 25 caracteres para não dar prolbema no excel
                if len(plan) > 25:
                    plan = plan[:25]   
                lista_planilhas.append(plan)

        wb = load_workbook(path_planilha_principal)      
        
        lista_arq_descompactados = [] 
        for p in os.listdir(pasta_descompactados):
            lista_arq_descompactados.append(os.path.join(pasta_descompactados, p))                    
        
        for arquivo_desc, aba in zip(lista_arq_descompactados, lista_planilhas):           
            wb_origem = load_workbook(arquivo_desc)
            
        # Pega a primeira aba ativa da planilha de origem
            ws_origem = wb_origem.active
            
            # Cria uma nova aba na planilha de destino com o nome da lista
            ws_destino = wb.create_sheet(title=aba)
            
            # Copia as dimensões das colunas (largura)
            for col, col_dim in ws_origem.column_dimensions.items():
                ws_destino.column_dimensions[col].width = col_dim.width
                
            # Copia as dimensões das linhas (altura)
            for row, row_dim in ws_origem.row_dimensions.items():
                ws_destino.row_dimensions[row].height = row_dim.height

            # Copia os dados e os estilos célula por célula
            for row in ws_origem.iter_rows():
                for cell in row:
                    # Cria a nova célula no destino
                    nova_celula = ws_destino.cell(row=cell.row, column=cell.column, value=cell.value)
                    
                    # Copia os estilos (fontes, cores, bordas, alinhamento)
                    if cell.has_style:
                        nova_celula.font = copy.copy(cell.font)
                        nova_celula.fill = copy.copy(cell.fill)
                        nova_celula.border = copy.copy(cell.border)
                        nova_celula.alignment = copy.copy(cell.alignment)
                        nova_celula.number_format = copy.copy(cell.number_format)
                        
                wb_origem.close()                

        # 4. Salva as alterações na planilha existente
        wb.save(path_planilha_principal)
        
        # DELETA A ABA PLANILHAFOLHA que é a duplicata da 412
        deleta_aba = wb[nome_plan_original]
        wb.remove(deleta_aba) 
        
        wb.save(path_planilha_principal) 
        wb.close()         
                    
        periodo = uploaded_file.name.replace("ENC_ Movimentação planilhas lançamento folha ","").replace(".zip", "")
        
        nome_planilha_final = f'{nome_plan_original} {periodo}.xlsx'
        
        for arq in os.listdir(pasta_descompactados):
            if arq == planilha_principal:                
                s = os.path.join(pasta_descompactados, arq)
                d = os.path.join(pasta_atual, nome_planilha_final)
                shutil.move(s,d)
            else:
                continue   
            
        #EXCLUI A PASTA DESCOMPACTADA PARA ATUALIZAÇAO
        if os.path.exists(pasta_descompactados):
            shutil.rmtree(pasta_descompactados)
            
        st.success(f'Planilha liberada:    {nome_planilha_final}')
        time.sleep(.5)
        st.balloons()            
            
    else:
        st.error("Por favor selecione a pasta para contiunar.")
            
              
        
       
        
            
